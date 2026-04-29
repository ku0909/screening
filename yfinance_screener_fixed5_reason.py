import pandas as pd
import yfinance as yf
import time
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 設定
# =========================
INPUT_CODES = "sample_codes_japan.csv"   # 取得したい銘柄コード一覧
NAME_CSV = "data_j(Sheet1).csv"          # 日本語銘柄名CSV
OUTPUT = "result.xlsx"
SLEEP = 0.5


def normalize_code(value):
    """銘柄コードを4桁文字列に統一する"""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(".T", "").replace(".t", "")
    return s.zfill(4)[:4]


def read_codes(path):
    """sample_codes_japan.csv から銘柄コードを読む"""
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    col = df.columns[0]
    return df[col].dropna().map(normalize_code).tolist()


def load_japanese_master():
    """
    data_j(Sheet1).csv から日本語銘柄名・市場・業種を読む。
    想定列:
      コード
      銘柄名
      市場・商品区分
      33業種区分
      17業種区分
      規模区分
    """
    df = pd.read_csv(NAME_CSV, dtype=str, encoding="utf-8-sig")
    df.columns = [str(c).strip() for c in df.columns]

    if "コード" not in df.columns or "銘柄名" not in df.columns:
        raise RuntimeError(f"{NAME_CSV} に必要列「コード」「銘柄名」がありません。列一覧: {df.columns.tolist()}")

    keep_cols = ["コード", "銘柄名"]
    for c in ["市場・商品区分", "33業種区分", "17業種区分", "規模区分"]:
        if c in df.columns:
            keep_cols.append(c)

    master = df[keep_cols].copy()
    master["コード"] = master["コード"].map(normalize_code)

    master = master.rename(columns={
        "銘柄名": "銘柄名_日本語",
        "市場・商品区分": "市場",
        "33業種区分": "33業種",
        "17業種区分": "17業種",
        "規模区分": "規模区分",
    })
    master = master.drop_duplicates(subset=["コード"], keep="first")
    return master


def calc_deviation(price, ma):
    """移動平均乖離率 = (現在値 - 移動平均) / 移動平均 * 100"""
    try:
        if price is None or ma is None or pd.isna(price) or pd.isna(ma) or ma == 0:
            return None
        return (price - ma) / ma * 100
    except Exception:
        return None


def fetch_price_history(ticker):
    """
    2年分の株価履歴から移動平均乖離率・52週高値安値を計算。
    """
    try:
        hist = yf.Ticker(ticker).history(period="2y", auto_adjust=False)
        if hist is None or hist.empty or "Close" not in hist.columns:
            return {}

        close = hist["Close"].dropna()
        if close.empty:
            return {}

        latest_price = float(close.iloc[-1])
        ma5 = close.tail(5).mean() if len(close) >= 5 else None
        ma25 = close.tail(25).mean() if len(close) >= 25 else None
        ma75 = close.tail(75).mean() if len(close) >= 75 else None
        ma200 = close.tail(200).mean() if len(close) >= 200 else None
        last_52w = close.tail(252) if len(close) >= 252 else close

        return {
            "履歴株価": latest_price,
            "5日線カイリ率(％)": calc_deviation(latest_price, ma5),
            "25日線カイリ率(％)": calc_deviation(latest_price, ma25),
            "75日線カイリ率(％)": calc_deviation(latest_price, ma75),
            "200日線カイリ率(％)": calc_deviation(latest_price, ma200),
            "52週高値": float(last_52w.max()) if not last_52w.empty else None,
            "52週安値": float(last_52w.min()) if not last_52w.empty else None,
        }
    except Exception:
        return {}


def fetch(code):
    """yfinanceから1銘柄分のデータを取得"""
    ticker = f"{code}.T"
    t = yf.Ticker(ticker)

    row = {
        "コード": code,
        "銘柄名": "",
        "市場": "",
        "33業種": "",
        "17業種": "",
        "規模区分": "",
        "Ticker": ticker,
        "株価": None,
        "5日線カイリ率(％)": None,
        "25日線カイリ率(％)": None,
        "75日線カイリ率(％)": None,
        "200日線カイリ率(％)": None,
        "52週高値": None,
        "52週安値": None,
        "配当利回り_%": None,
        "予想年間配当": None,
        "PER\n（10倍以下）": None,
        "PBR\n（1.0倍以下）": None,
        "PER×PBR\n（15倍以下）": None,
        "EV/EBITDA（10倍以下）": None,
        "時価総額\n(百万円)": None,
        "時価総額_億円": None,
        "取得メモ": "",
    }

    try:
        info = t.info or {}
        row["銘柄名"] = info.get("longName") or info.get("shortName") or ""

        hist_data = fetch_price_history(ticker)
        row.update(hist_data)

        price = (
            info.get("currentPrice")
            or info.get("regularMarketPrice")
            or info.get("previousClose")
            or row.get("履歴株価")
        )
        row["株価"] = price

        dy = info.get("dividendYield") or info.get("trailingAnnualDividendYield")
        if dy is not None:
            dy = float(dy)
            row["配当利回り_%"] = dy * 100 if dy <= 1 else dy

        annual_div = info.get("trailingAnnualDividendRate") or info.get("dividendRate")
        if annual_div is not None:
            row["予想年間配当"] = annual_div

        per = info.get("trailingPE") or info.get("forwardPE")
        pbr = info.get("priceToBook")
        row["PER\n（10倍以下）"] = per
        row["PBR\n（1.0倍以下）"] = pbr

        try:
            if per is not None and pbr is not None:
                row["PER×PBR\n（15倍以下）"] = float(per) * float(pbr)
        except Exception:
            pass

        row["EV/EBITDA（10倍以下）"] = info.get("enterpriseToEbitda")

        mc = info.get("marketCap")
        if mc is not None:
            row["時価総額\n(百万円)"] = float(mc) / 1_000_000
            row["時価総額_億円"] = float(mc) / 100_000_000

    except Exception as e:
        row["取得メモ"] = str(e)

    time.sleep(SLEEP)
    return row


def judge(score_value):
    """スコアから判定を返す。上書きバグ防止のためif文で判定。"""
    if score_value >= 90:
        return "A:本命候補"
    elif score_value >= 70:
        return "B:強い候補"
    elif score_value >= 50:
        return "C:監視"
    else:
        return "D:見送り"


def score(df):
    """スコアと判定を作成"""
    df["スコア"] = 0

    # 配当
    df.loc[df["配当利回り_%"] >= 3.5, "スコア"] += 30
    df.loc[df["配当利回り_%"] >= 4.5, "スコア"] += 10

    # バリュー
    df.loc[df["PER\n（10倍以下）"] <= 10, "スコア"] += 25
    df.loc[df["PBR\n（1.0倍以下）"] <= 1.0, "スコア"] += 25
    df.loc[df["PER×PBR\n（15倍以下）"] <= 15, "スコア"] += 20
    df.loc[df["EV/EBITDA（10倍以下）"] <= 10, "スコア"] += 15

    # 安全性
    df.loc[df["時価総額\n(百万円)"] >= 100000, "スコア"] += 10

    # タイミング
    df.loc[df["5日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["25日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["75日線カイリ率(％)"] < 0, "スコア"] += 5

    # 価格位置
    df.loc[df["株価"] < df["52週高値"] * 0.8, "スコア"] += 10
    df.loc[df["株価"] > df["52週安値"] * 1.1, "スコア"] += 5

    df["判定"] = df["スコア"].apply(judge)
    return df.sort_values(["スコア", "配当利回り_%"], ascending=[False, False])


def make_reason(row):
    """高評価銘柄向けの購入推奨理由を作成"""
    if row.get("スコア", 0) < 50:
        return ""

    reasons = []

    # 配当
    dy = row.get("配当利回り_%")
    if pd.notna(dy):
        if dy >= 4.5:
            reasons.append("高配当4.5%以上")
        elif dy >= 3.5:
            reasons.append("高配当3.5%以上")

    # 割安
    per = row.get("PER\n（10倍以下）")
    pbr = row.get("PBR\n（1.0倍以下）")
    per_pbr = row.get("PER×PBR\n（15倍以下）")
    ev_ebitda = row.get("EV/EBITDA（10倍以下）")

    if pd.notna(per) and per <= 10:
        reasons.append("PER10倍以下で割安")
    if pd.notna(pbr) and pbr <= 1.0:
        reasons.append("PBR1倍以下")
    if pd.notna(per_pbr) and per_pbr <= 15:
        reasons.append("PER×PBRが15以下")
    if pd.notna(ev_ebitda) and ev_ebitda <= 10:
        reasons.append("EV/EBITDA10倍以下")

    # 安全性
    market_cap = row.get("時価総額\n(百万円)")
    if pd.notna(market_cap) and market_cap >= 100000:
        reasons.append("時価総額1000億円以上")

    # タイミング
    d25 = row.get("25日線カイリ率(％)")
    d75 = row.get("75日線カイリ率(％)")
    if pd.notna(d25) and d25 < 0:
        reasons.append("25日線下回りで押し目")
    if pd.notna(d75) and d75 < 0:
        reasons.append("75日線下回りで安値圏")

    # 52週位置
    price = row.get("株価")
    high52 = row.get("52週高値")
    low52 = row.get("52週安値")
    if pd.notna(price) and pd.notna(high52) and price < high52 * 0.8:
        reasons.append("52週高値から20%以上下落")
    if pd.notna(price) and pd.notna(low52) and price > low52 * 1.1:
        reasons.append("52週安値から反発")

    return " / ".join(reasons)


def apply_excel_format(writer, df):
    """1行目フィルタ・色・列幅・固定を設定"""
    wb = writer.book

    header_fill = PatternFill("solid", fgColor="9DC3E6")
    header_font = Font(bold=True)
    warning_fill = PatternFill("solid", fgColor="E4DFEC")
    good_fill = PatternFill("solid", fgColor="E2F0D9")
    buy_fill = PatternFill("solid", fgColor="C6E0B4")
    watch_fill = PatternFill("solid", fgColor="FFF2CC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # ヘッダー
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 条件項目っぽい列に薄紫
        for cell in ws[1]:
            if any(x in str(cell.value) for x in ["10倍以下", "1.0倍以下", "15倍以下", "EV/EBITDA", "時価総額"]):
                cell.fill = warning_fill

        # 列幅
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)

        ws.row_dimensions[1].height = 36

        # 判定で色付け
        header_values = [c.value for c in ws[1]]
        if "判定" in header_values:
            verdict_col = header_values.index("判定") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                verdict = row[verdict_col - 1].value
                if verdict == "A:本命候補":
                    for c in row:
                        c.fill = buy_fill
                elif verdict == "B:強い候補":
                    for c in row:
                        c.fill = good_fill
                elif verdict == "C:監視":
                    for c in row:
                        c.fill = watch_fill


def main():
    codes = read_codes(INPUT_CODES)
    print(f"{len(codes)}銘柄取得開始")

    rows = []
    for i, code in enumerate(codes, start=1):
        print(f"{i}/{len(codes)}: {code}.T")
        rows.append(fetch(code))

    df = pd.DataFrame(rows)
    df["コード"] = df["コード"].map(normalize_code)

    # 日本語銘柄名CSVを結合
    try:
        master = load_japanese_master()
        df = df.merge(master, on="コード", how="left", suffixes=("", "_master"))

        df["銘柄名"] = df["銘柄名_日本語"].where(
            df["銘柄名_日本語"].notna() & (df["銘柄名_日本語"].astype(str).str.strip() != ""),
            df["銘柄名"]
        )
        df = df.drop(columns=["銘柄名_日本語"])

        for col in ["市場", "33業種", "17業種", "規模区分"]:
            master_col = f"{col}_master"
            if master_col in df.columns:
                df[col] = df[master_col].where(df[master_col].notna(), df[col])
                df = df.drop(columns=[master_col])

        print("日本語銘柄名CSV適用OK")
    except Exception as e:
        print("日本語銘柄名CSVエラー:", e)

    # スコア・判定・購入理由を作成
    df = score(df)
    df["購入推奨理由"] = df.apply(make_reason, axis=1)

    # 小数は第2位までに丸める
    df = df.round(2)

    preferred_cols = [
        "コード", "銘柄名", "市場", "33業種", "17業種", "規模区分", "Ticker",
        "株価",
        "配当利回り_%",
        "予想年間配当",
        "PER\n（10倍以下）",
        "PBR\n（1.0倍以下）",
        "PER×PBR\n（15倍以下）",
        "EV/EBITDA（10倍以下）",
        "5日線カイリ率(％)",
        "25日線カイリ率(％)",
        "75日線カイリ率(％)",
        "200日線カイリ率(％)",
        "52週高値",
        "52週安値",
        "時価総額\n(百万円)",
        "スコア",
        "判定",
        "購入推奨理由",
        "取得メモ",
    ]
    cols = [c for c in preferred_cols if c in df.columns] + [c for c in df.columns if c not in preferred_cols]
    df = df[cols]

    output_path = Path(OUTPUT).resolve()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df[df["判定"] == "A:本命候補"].to_excel(writer, index=False, sheet_name="A本命候補")
        df[df["判定"].isin(["A:本命候補", "B:強い候補"])].to_excel(writer, index=False, sheet_name="AB候補")
        df.head(100).to_excel(writer, index=False, sheet_name="ランキング")
        df.to_excel(writer, index=False, sheet_name="全銘柄")
        apply_excel_format(writer, df)

    print("完了:", output_path)


if __name__ == "__main__":
    main()
