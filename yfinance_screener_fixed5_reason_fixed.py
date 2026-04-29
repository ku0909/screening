# -*- coding: utf-8 -*-
"""
Google Colab用：高配当株スクリーナー Drive固定フォルダ版

【事前準備】
Google Driveのマイドライブに「株スクリーナー」フォルダを作り、以下2ファイルを入れてください。
- sample_codes_japan.csv
- data_j(Sheet1).csv

【Colabでの実行】
1. 最初のセルで必要ライブラリを入れる
   !pip install -q yfinance openpyxl pandas

2. このPythonファイルをアップロードする、またはDriveに置く

3. 実行
   !python high_dividend_screener_drive_colab.py

【出力】
Google Driveの「株スクリーナー」フォルダに以下が出ます。
- result.html  スマホで見やすいHTML
- result.xlsx  Excel版
"""

import time
from pathlib import Path
import html

import pandas as pd
import yfinance as yf
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =========================
# Google Drive固定パス
# =========================
try:
    from google.colab import drive
    drive.mount("/content/drive")
    BASE_DIR = Path("/content/drive/MyDrive/株スクリーナー")
except Exception:
    BASE_DIR = Path(__file__).resolve().parent


INPUT_CODES = BASE_DIR / "sample_codes_japan.csv"
NAME_CSV = BASE_DIR / "data_j(Sheet1).csv"
OUTPUT_HTML = BASE_DIR / "result.html"
OUTPUT_XLSX = BASE_DIR / "result.xlsx"

SLEEP = 0.5


def normalize_code(value):
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(".T", "").replace(".t", "")
    return s.zfill(4)[:4]


def read_codes(path):
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
    col = df.columns[0]
    return df[col].dropna().map(normalize_code).tolist()


def load_japanese_master():
    df = pd.read_csv(NAME_CSV, dtype=str, encoding="utf-8-sig")
    df.columns = [str(c).strip() for c in df.columns]

    if "コード" not in df.columns or "銘柄名" not in df.columns:
        raise RuntimeError(f"{NAME_CSV} に必要列「コード」「銘柄名」がありません。列一覧: {df.columns.tolist()}")

    keep_cols = ["コード", "銘柄名"]
    for c in ["市場・商品区分", "33業種区分", "17業種区分", "規模区分"]:
        if c in df.columns:
            keep_cols.append(c)

    master = df[keep_cols].copy()
    master["コード"] = master["コード"].map(normalize_code)
    master = master.rename(columns={
        "銘柄名": "銘柄名_日本語",
        "市場・商品区分": "市場",
        "33業種区分": "33業種",
        "17業種区分": "17業種",
        "規模区分": "規模区分",
    })
    return master.drop_duplicates(subset=["コード"], keep="first")


def calc_deviation(price, ma):
    try:
        if price is None or ma is None or pd.isna(price) or pd.isna(ma) or ma == 0:
            return None
        return (price - ma) / ma * 100
    except Exception:
        return None


def fetch_price_history(ticker):
    try:
        hist = yf.Ticker(ticker).history(period="2y", auto_adjust=False)
        if hist is None or hist.empty or "Close" not in hist.columns:
            return {}

        close = hist["Close"].dropna()
        if close.empty:
            return {}

        latest_price = float(close.iloc[-1])
        ma5 = close.tail(5).mean() if len(close) >= 5 else None
        ma25 = close.tail(25).mean() if len(close) >= 25 else None
        ma75 = close.tail(75).mean() if len(close) >= 75 else None
        ma200 = close.tail(200).mean() if len(close) >= 200 else None
        last_52w = close.tail(252) if len(close) >= 252 else close

        return {
            "履歴株価": latest_price,
            "5日線カイリ率(％)": calc_deviation(latest_price, ma5),
            "25日線カイリ率(％)": calc_deviation(latest_price, ma25),
            "75日線カイリ率(％)": calc_deviation(latest_price, ma75),
            "200日線カイリ率(％)": calc_deviation(latest_price, ma200),
            "52週高値": float(last_52w.max()) if not last_52w.empty else None,
            "52週安値": float(last_52w.min()) if not last_52w.empty else None,
        }
    except Exception:
        return {}


def fetch(code):
    ticker = f"{code}.T"
    t = yf.Ticker(ticker)

    row = {
        "コード": code,
        "銘柄名": "",
        "市場": "",
        "33業種": "",
        "17業種": "",
        "規模区分": "",
        "Ticker": ticker,
        "株価": None,
        "5日線カイリ率(％)": None,
        "25日線カイリ率(％)": None,
        "75日線カイリ率(％)": None,
        "200日線カイリ率(％)": None,
        "52週高値": None,
        "52週安値": None,
        "配当利回り_%": None,
        "予想年間配当": None,
        "PER\n（10倍以下）": None,
        "PBR\n（1.0倍以下）": None,
        "PER×PBR\n（15倍以下）": None,
        "EV/EBITDA（10倍以下）": None,
        "時価総額\n(百万円)": None,
        "時価総額_億円": None,
        "売上成長率_%": None,
        "利益成長率_%": None,
        "負債比率_%": None,
        "取得メモ": "",
    }

    try:
        info = t.info or {}
        row["銘柄名"] = info.get("longName") or info.get("shortName") or ""
        row.update(fetch_price_history(ticker))

        price = (
            info.get("currentPrice")
            or info.get("regularMarketPrice")
            or info.get("previousClose")
            or row.get("履歴株価")
        )
        row["株価"] = price

        dy = info.get("dividendYield") or info.get("trailingAnnualDividendYield")
        if dy is not None:
            dy = float(dy)
            row["配当利回り_%"] = dy * 100 if dy <= 1 else dy

        annual_div = info.get("trailingAnnualDividendRate") or info.get("dividendRate")
        if annual_div is not None:
            row["予想年間配当"] = annual_div

        per = info.get("trailingPE") or info.get("forwardPE")
        pbr = info.get("priceToBook")
        row["PER\n（10倍以下）"] = per
        row["PBR\n（1.0倍以下）"] = pbr

        if per is not None and pbr is not None:
            try:
                row["PER×PBR\n（15倍以下）"] = float(per) * float(pbr)
            except Exception:
                pass

        row["EV/EBITDA（10倍以下）"] = info.get("enterpriseToEbitda")

        mc = info.get("marketCap")
        if mc is not None:
            row["時価総額\n(百万円)"] = float(mc) / 1_000_000
            row["時価総額_億円"] = float(mc) / 100_000_000

        rg = info.get("revenueGrowth")
        eg = info.get("earningsGrowth")
        de = info.get("debtToEquity")
        if rg is not None:
            row["売上成長率_%"] = float(rg) * 100
        if eg is not None:
            row["利益成長率_%"] = float(eg) * 100
        if de is not None:
            row["負債比率_%"] = float(de)

    except Exception as e:
        row["取得メモ"] = str(e)

    time.sleep(SLEEP)
    return row


def judge(score_value):
    if score_value >= 90:
        return "A:本命候補"
    elif score_value >= 70:
        return "B:強い候補"
    elif score_value >= 50:
        return "C:監視"
    else:
        return "D:見送り"


def score_df(df):
    df = df.copy()
    df["スコア"] = 0
    df["危険減点"] = 0

    df.loc[df["配当利回り_%"] >= 3.5, "スコア"] += 30
    df.loc[df["配当利回り_%"] >= 4.5, "スコア"] += 10
    df.loc[df["PER\n（10倍以下）"] <= 10, "スコア"] += 25
    df.loc[df["PBR\n（1.0倍以下）"] <= 1.0, "スコア"] += 25
    df.loc[df["PER×PBR\n（15倍以下）"] <= 15, "スコア"] += 20
    df.loc[df["EV/EBITDA（10倍以下）"] <= 10, "スコア"] += 15
    df.loc[df["時価総額\n(百万円)"] >= 100000, "スコア"] += 10
    df.loc[df["時価総額\n(百万円)"] >= 500000, "スコア"] += 5
    df.loc[df["5日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["25日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["75日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["株価"] < df["52週高値"] * 0.8, "スコア"] += 10
    df.loc[df["株価"] > df["52週安値"] * 1.1, "スコア"] += 5

    df.loc[df["配当利回り_%"] > 6, "危険減点"] += 25
    df.loc[df["配当利回り_%"] > 7, "危険減点"] += 20
    df.loc[df["株価"] < df["52週高値"] * 0.6, "危険減点"] += 30
    df.loc[df["75日線カイリ率(％)"] < -10, "危険減点"] += 20
    df.loc[df["PER\n（10倍以下）"] < 5, "危険減点"] += 10
    df.loc[df["売上成長率_%"] < 0, "危険減点"] += 20
    df.loc[df["利益成長率_%"] < 0, "危険減点"] += 20
    df.loc[df["負債比率_%"] > 150, "危険減点"] += 15

    if "33業種" in df.columns:
        df.loc[df["33業種"].isin(["鉄鋼", "海運業", "鉱業"]), "危険減点"] += 10

    df.loc[
        (df["配当利回り_%"] > 4) & (df["PER\n（10倍以下）"] > 20),
        "危険減点"
    ] += 20

    df["総合スコア"] = df["スコア"] - df["危険減点"]
    df["判定"] = df["総合スコア"].apply(judge)

    return df.sort_values(["総合スコア", "配当利回り_%"], ascending=[False, False])


def make_reason(row):
    if row.get("総合スコア", 0) < 50:
        return ""

    reasons = []

    dy = row.get("配当利回り_%")
    if pd.notna(dy):
        if dy >= 4.5:
            reasons.append("高配当4.5%以上")
        elif dy >= 3.5:
            reasons.append("高配当3.5%以上")
        if dy > 6:
            reasons.append("⚠配当高すぎ（減配リスク）")

    per = row.get("PER\n（10倍以下）")
    pbr = row.get("PBR\n（1.0倍以下）")
    per_pbr = row.get("PER×PBR\n（15倍以下）")
    ev_ebitda = row.get("EV/EBITDA（10倍以下）")

    if pd.notna(per) and per <= 10:
        reasons.append("PER10倍以下で割安")
    if pd.notna(per) and per < 5:
        reasons.append("⚠PER低すぎ（要確認）")
    if pd.notna(pbr) and pbr <= 1:
        reasons.append("PBR1倍以下")
    if pd.notna(per_pbr) and per_pbr <= 15:
        reasons.append("PER×PBRが15以下")
    if pd.notna(ev_ebitda) and ev_ebitda <= 10:
        reasons.append("EV/EBITDA10倍以下")

    market_cap = row.get("時価総額\n(百万円)")
    if pd.notna(market_cap) and market_cap >= 100000:
        reasons.append("時価総額1000億円以上")

    d25 = row.get("25日線カイリ率(％)")
    d75 = row.get("75日線カイリ率(％)")
    if pd.notna(d25) and d25 < 0:
        reasons.append("25日線下回りで押し目")
    if pd.notna(d75) and d75 < 0:
        reasons.append("75日線下回り")
    if pd.notna(d75) and d75 < -10:
        reasons.append("⚠下落トレンド強い")

    price = row.get("株価")
    high52 = row.get("52週高値")
    low52 = row.get("52週安値")
    if pd.notna(price) and pd.notna(high52) and price < high52 * 0.8:
        reasons.append("52週高値から20%以上下落")
    if pd.notna(price) and pd.notna(high52) and price < high52 * 0.6:
        reasons.append("⚠下がりすぎ")
    if pd.notna(price) and pd.notna(low52) and price > low52 * 1.1:
        reasons.append("52週安値から反発")

    if pd.notna(row.get("売上成長率_%")) and row.get("売上成長率_%") < 0:
        reasons.append("⚠売上減少")
    if pd.notna(row.get("利益成長率_%")) and row.get("利益成長率_%") < 0:
        reasons.append("⚠利益減少")
    if pd.notna(row.get("負債比率_%")) and row.get("負債比率_%") > 150:
        reasons.append("⚠負債比率高い")

    if row.get("33業種") in ["鉄鋼", "海運業", "鉱業"]:
        reasons.append("⚠景気敏感業種")

    return " / ".join(reasons)


def reorder_columns(df):
    preferred_cols = [
        "コード", "銘柄名", "市場", "33業種", "17業種", "規模区分", "Ticker",
        "株価", "配当利回り_%", "予想年間配当",
        "PER\n（10倍以下）", "PBR\n（1.0倍以下）", "PER×PBR\n（15倍以下）",
        "EV/EBITDA（10倍以下）",
        "5日線カイリ率(％)", "25日線カイリ率(％)", "75日線カイリ率(％)", "200日線カイリ率(％)",
        "52週高値", "52週安値",
        "時価総額\n(百万円)", "売上成長率_%", "利益成長率_%", "負債比率_%",
        "スコア", "危険減点", "総合スコア", "判定", "購入推奨理由", "取得メモ"
    ]
    cols = [c for c in preferred_cols if c in df.columns] + [c for c in df.columns if c not in preferred_cols]
    return df[cols]


def save_excel(df):
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df[df["判定"] == "A:本命候補"].to_excel(writer, index=False, sheet_name="A本命候補")
        df[df["判定"].isin(["A:本命候補", "B:強い候補"])].to_excel(writer, index=False, sheet_name="AB候補")
        df.head(100).to_excel(writer, index=False, sheet_name="ランキング")
        df.to_excel(writer, index=False, sheet_name="全銘柄")

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="9DC3E6")
        header_font = Font(bold=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.row_dimensions[1].height = 36

            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                max_len = 0
                for cell in col[:200]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)


def fmt(v, suffix=""):
    if pd.isna(v):
        return "-"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}{suffix}"
    return html.escape(str(v))


def make_html_report(df):
    top = df[df["判定"].isin(["A:本命候補", "B:強い候補"])].copy()
    if top.empty:
        top = df.head(20).copy()

    now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    cards = []

    for _, r in top.iterrows():
        verdict = str(r.get("判定", ""))
        cls = "rank-a" if verdict.startswith("A") else "rank-b" if verdict.startswith("B") else "rank-c"

        reason = str(r.get("購入推奨理由", "") or "")
        reason_html = "".join(f"<li>{html.escape(x.strip())}</li>" for x in reason.split("/") if x.strip())

        # f-string内ではバックスラッシュを含む列名を直接書けないため、先に取り出す
        per_value = fmt(r.get("PER\n（10倍以下）"))
        pbr_value = fmt(r.get("PBR\n（1.0倍以下）"))
        per_pbr_value = fmt(r.get("PER×PBR\n（15倍以下）"))

        cards.append(f"""
        <section class="card {cls}">
          <div class="topline">
            <div>
              <div class="code">{html.escape(str(r.get('コード', '')))} / {html.escape(str(r.get('Ticker', '')))}</div>
              <h2>{html.escape(str(r.get('銘柄名', '')))}</h2>
              <div class="sub">{html.escape(str(r.get('市場','')))}｜{html.escape(str(r.get('33業種','')))}｜{html.escape(str(r.get('規模区分','')))}</div>
            </div>
            <div class="badge">{html.escape(verdict)}</div>
          </div>

          <div class="scorebox">
            <div><span>総合スコア</span><strong>{fmt(r.get('総合スコア'))}</strong></div>
            <div><span>加点</span><strong>{fmt(r.get('スコア'))}</strong></div>
            <div><span>危険減点</span><strong>{fmt(r.get('危険減点'))}</strong></div>
          </div>

          <div class="grid">
            <div><span>株価</span><b>{fmt(r.get('株価'))}</b></div>
            <div><span>配当利回り</span><b>{fmt(r.get('配当利回り_%'), '%')}</b></div>
            <div><span>PER</span><b>{per_value}</b></div>
            <div><span>PBR</span><b>{pbr_value}</b></div>
            <div><span>PER×PBR</span><b>{per_pbr_value}</b></div>
            <div><span>EV/EBITDA</span><b>{fmt(r.get('EV/EBITDA（10倍以下）'))}</b></div>
            <div><span>25日乖離</span><b>{fmt(r.get('25日線カイリ率(％)'), '%')}</b></div>
            <div><span>52週高値</span><b>{fmt(r.get('52週高値'))}</b></div>
          </div>

          <div class="reason">
            <h3>理由・注意点</h3>
            <ul>{reason_html or '<li>-</li>'}</ul>
          </div>
        </section>
        """)

    html_doc = f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>高配当株スクリーニング結果</title>
  <style>
    body {{
      margin: 0;
      background: #f5f6f8;
      font-family: -apple-system, BlinkMacSystemFont, "Helvetica Neue", Arial, sans-serif;
      color: #1f2937;
    }}
    header {{
      position: sticky;
      top: 0;
      background: #111827;
      color: white;
      padding: 14px 16px;
      z-index: 10;
    }}
    header h1 {{
      font-size: 20px;
      margin: 0 0 4px;
    }}
    header p {{
      margin: 0;
      font-size: 12px;
      opacity: .8;
    }}
    main {{
      max-width: 760px;
      margin: 0 auto;
      padding: 12px;
    }}
    .summary {{
      background: white;
      border-radius: 16px;
      padding: 14px;
      margin-bottom: 12px;
      box-shadow: 0 2px 10px rgba(0,0,0,.06);
    }}
    .summary strong {{
      font-size: 22px;
    }}
    .card {{
      background: white;
      border-radius: 18px;
      padding: 16px;
      margin: 12px 0;
      box-shadow: 0 4px 18px rgba(0,0,0,.08);
      border-left: 7px solid #9ca3af;
    }}
    .rank-a {{ border-left-color: #16a34a; }}
    .rank-b {{ border-left-color: #2563eb; }}
    .rank-c {{ border-left-color: #f59e0b; }}
    .topline {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      align-items: flex-start;
    }}
    .code {{ font-size: 12px; color: #6b7280; }}
    h2 {{ margin: 3px 0; font-size: 21px; }}
    .sub {{ font-size: 12px; color: #6b7280; }}
    .badge {{
      white-space: nowrap;
      border-radius: 999px;
      padding: 7px 10px;
      font-size: 12px;
      font-weight: 700;
      background: #eef2ff;
      color: #1e3a8a;
    }}
    .scorebox {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 8px;
      margin: 14px 0;
    }}
    .scorebox div {{
      background: #f9fafb;
      border-radius: 12px;
      padding: 10px;
      text-align: center;
    }}
    .scorebox span {{
      display: block;
      font-size: 11px;
      color: #6b7280;
    }}
    .scorebox strong {{ font-size: 22px; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 8px;
    }}
    .grid div {{
      background: #f3f4f6;
      border-radius: 12px;
      padding: 9px 10px;
    }}
    .grid span {{
      display: block;
      font-size: 11px;
      color: #6b7280;
    }}
    .grid b {{ font-size: 16px; }}
    .reason {{
      margin-top: 12px;
      background: #fffbeb;
      border-radius: 12px;
      padding: 10px 12px;
    }}
    .reason h3 {{
      margin: 0 0 6px;
      font-size: 14px;
    }}
    .reason ul {{
      margin: 0;
      padding-left: 20px;
    }}
    .reason li {{
      margin: 3px 0;
      font-size: 14px;
    }}
  </style>
</head>
<body>
  <header>
    <h1>高配当株スクリーニング結果</h1>
    <p>作成日時: {now}</p>
  </header>
  <main>
    <div class="summary">
      表示対象: <strong>{len(top)}</strong>件 / 全銘柄: {len(df)}件<br>
      A候補: {(df['判定'] == 'A:本命候補').sum()}件　
      B候補: {(df['判定'] == 'B:強い候補').sum()}件
    </div>
    {''.join(cards)}
  </main>
</body>
</html>"""

    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_HTML.write_text(html_doc, encoding="utf-8")


def main():
    BASE_DIR.mkdir(parents=True, exist_ok=True)

    print("使用フォルダ:", BASE_DIR)
    print("銘柄コードCSV:", INPUT_CODES)
    print("銘柄名マスター:", NAME_CSV)

    if not INPUT_CODES.exists():
        raise FileNotFoundError(f"{INPUT_CODES} が見つかりません。Google Driveの「株スクリーナー」フォルダに置いてください。")

    if not NAME_CSV.exists():
        raise FileNotFoundError(f"{NAME_CSV} が見つかりません。Google Driveの「株スクリーナー」フォルダに置いてください。")

    codes = read_codes(INPUT_CODES)
    print(f"{len(codes)}銘柄取得開始")

    rows = []
    for i, code in enumerate(codes, start=1):
        print(f"{i}/{len(codes)}: {code}.T")
        rows.append(fetch(code))

    df = pd.DataFrame(rows)
    df["コード"] = df["コード"].map(normalize_code)

    try:
        master = load_japanese_master()
        df = df.merge(master, on="コード", how="left", suffixes=("", "_master"))

        df["銘柄名"] = df["銘柄名_日本語"].where(
            df["銘柄名_日本語"].notna() & (df["銘柄名_日本語"].astype(str).str.strip() != ""),
            df["銘柄名"]
        )
        df = df.drop(columns=["銘柄名_日本語"])

        for col in ["市場", "33業種", "17業種", "規模区分"]:
            master_col = f"{col}_master"
            if master_col in df.columns:
                df[col] = df[master_col].where(df[master_col].notna(), df[col])
                df = df.drop(columns=[master_col])

        print("日本語銘柄名CSV適用OK")
    except Exception as e:
        print("日本語銘柄名CSVエラー:", e)

    df = score_df(df)
    df["購入推奨理由"] = df.apply(make_reason, axis=1)
    df = df.round(2)
    df = reorder_columns(df)

    save_excel(df)
    make_html_report(df)

    print("完了:")
    print("HTML:", OUTPUT_HTML)
    print("Excel:", OUTPUT_XLSX)

    return df


if __name__ == "__main__":
    main()
