import pandas as pd
import yfinance as yf
import time
import html
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# =========================
# 設定
# =========================
INPUT_CODES = "sample_codes_japan.csv"   # 例外的に追加したい銘柄コード一覧
NAME_CSV = "data_j(Sheet1).csv"          # 日本語銘柄名CSV
OUTPUT = "result.xlsx"
OUTPUT_HTML = "result.html"
SLEEP = 0.5

# True:
#   data_j(Sheet1).csv のプライム銘柄を自動対象にする。
#   sample_codes_japan.csv にある銘柄は、プライム以外でも例外的に追加する。
AUTO_PICK_PRIME = True


# =========================
# 基本関数
# =========================
def normalize_code(value):
    """銘柄コードを4桁文字列に統一する"""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace(".T", "").replace(".t", "")
    return s.zfill(4)[:4]


def read_sample_codes(path):
    """sample_codes_japan.csv から例外追加銘柄コードを読む"""
    p = Path(path)
    if not p.exists():
        return []
    df = pd.read_csv(p, dtype=str, encoding="utf-8-sig")
    col = df.columns[0]
    codes = df[col].dropna().map(normalize_code).tolist()
    return list(dict.fromkeys(codes))


def load_japanese_master():
    """
    data_j(Sheet1).csv から日本語銘柄名・市場・業種を読む。
    想定列:
      コード
      銘柄名
      市場・商品区分
      33業種区分
      17業種区分
      規模区分
    """
    df = pd.read_csv(NAME_CSV, dtype=str, encoding="utf-8-sig")
    df.columns = [str(c).strip() for c in df.columns]

    if "コード" not in df.columns or "銘柄名" not in df.columns:
        raise RuntimeError(f"{NAME_CSV} に必要列「コード」「銘柄名」がありません。列一覧: {df.columns.tolist()}")

    keep_cols = ["コード", "銘柄名"]
    for c in ["市場・商品区分", "33業種区分", "17業種区分", "規模区分"]:
        if c in df.columns:
            keep_cols.append(c)

    master = df[keep_cols].copy()
    master["コード"] = master["コード"].map(normalize_code)

    master = master.rename(columns={
        "銘柄名": "銘柄名_日本語",
        "市場・商品区分": "市場",
        "33業種区分": "33業種",
        "17業種区分": "17業種",
        "規模区分": "規模区分",
    })
    master = master.drop_duplicates(subset=["コード"], keep="first")
    return master


def build_target_codes(master):
    """
    対象銘柄を作成。
    - プライム市場の銘柄を自動対象
    - sample_codes_japan.csv にある銘柄は例外追加
    """
    sample_codes = read_sample_codes(INPUT_CODES)

    if not AUTO_PICK_PRIME:
        return sample_codes

    if "市場" not in master.columns:
        raise RuntimeError("data_j(Sheet1).csv に市場列がありません。")

    prime = master[master["市場"].astype(str).str.contains("プライム", na=False)].copy()
    prime_codes = prime["コード"].dropna().map(normalize_code).tolist()

    # 順番を維持して重複除去
    codes = list(dict.fromkeys(prime_codes + sample_codes))

    print(f"プライム対象: {len(set(prime_codes))}銘柄")
    print(f"sample_codes_japan.csv 登録銘柄: {len(set(sample_codes))}銘柄")
    print(f"合計対象: {len(codes)}銘柄")

    return codes


def to_float(value):
    """単一値を安全にfloat化"""
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def calc_deviation(price, ma):
    """移動平均乖離率 = (現在値 - 移動平均) / 移動平均 * 100"""
    try:
        if price is None or ma is None or pd.isna(price) or pd.isna(ma) or ma == 0:
            return None
        return (price - ma) / ma * 100
    except Exception:
        return None


def fetch_price_history(ticker):
    """2年分の株価履歴から移動平均乖離率・52週高値安値を計算"""
    try:
        hist = yf.Ticker(ticker).history(period="2y", auto_adjust=False)
        if hist is None or hist.empty or "Close" not in hist.columns:
            return {}

        close = hist["Close"].dropna()
        if close.empty:
            return {}

        latest_price = float(close.iloc[-1])
        ma5 = close.tail(5).mean() if len(close) >= 5 else None
        ma25 = close.tail(25).mean() if len(close) >= 25 else None
        ma75 = close.tail(75).mean() if len(close) >= 75 else None
        ma200 = close.tail(200).mean() if len(close) >= 200 else None
        last_52w = close.tail(252) if len(close) >= 252 else close

        return {
            "履歴株価": latest_price,
            "5日線カイリ率(％)": calc_deviation(latest_price, ma5),
            "25日線カイリ率(％)": calc_deviation(latest_price, ma25),
            "75日線カイリ率(％)": calc_deviation(latest_price, ma75),
            "200日線カイリ率(％)": calc_deviation(latest_price, ma200),
            "52週高値": float(last_52w.max()) if not last_52w.empty else None,
            "52週安値": float(last_52w.min()) if not last_52w.empty else None,
        }
    except Exception:
        return {}


def fetch(code):
    """yfinanceから1銘柄分のデータを取得"""
    ticker = f"{code}.T"
    t = yf.Ticker(ticker)

    row = {
        "コード": code,
        "銘柄名": "",
        "市場": "",
        "33業種": "",
        "17業種": "",
        "規模区分": "",
        "Ticker": ticker,
        "株価": None,
        "5日線カイリ率(％)": None,
        "25日線カイリ率(％)": None,
        "75日線カイリ率(％)": None,
        "200日線カイリ率(％)": None,
        "52週高値": None,
        "52週安値": None,
        "年間配当_yfinance": None,
        "配当利回り_%": None,
        "予想年間配当": None,
        "PER\n（10倍以下）": None,
        "PBR\n（1.0倍以下）": None,
        "PER×PBR\n（15倍以下）": None,
        "EV/EBITDA（10倍以下）": None,
        "時価総額\n(百万円)": None,
        "時価総額_億円": None,
        "売上成長率_%": None,
        "利益成長率_%": None,
        "負債比率_%": None,
        "取得メモ": "",
    }

    try:
        info = t.info or {}
        row["銘柄名"] = info.get("longName") or info.get("shortName") or ""

        hist_data = fetch_price_history(ticker)
        row.update(hist_data)

        price = (
            info.get("currentPrice")
            or info.get("regularMarketPrice")
            or info.get("previousClose")
            or row.get("履歴株価")
        )
        price = to_float(price)
        row["株価"] = price

        # 年間配当は dividendRate を優先。なければ trailingAnnualDividendRate
        # 配当利回りは yfinance の dividendYield を使わず、年間配当 ÷ 株価で再計算する
        annual_div = info.get("dividendRate")
        if annual_div is None:
            annual_div = info.get("trailingAnnualDividendRate")
        annual_div = to_float(annual_div)

        if annual_div is not None:
            row["年間配当_yfinance"] = annual_div
            row["予想年間配当"] = annual_div

        if price is not None and price > 0 and annual_div is not None:
            row["配当利回り_%"] = annual_div / price * 100

        # PERは forwardPE を優先。なければ trailingPE。表示は採点用PERのみ。
        forward_pe = to_float(info.get("forwardPE"))
        trailing_pe = to_float(info.get("trailingPE"))
        score_per = forward_pe if forward_pe is not None else trailing_pe
        row["PER\n（10倍以下）"] = score_per

        pbr = to_float(info.get("priceToBook"))
        row["PBR\n（1.0倍以下）"] = pbr

        try:
            if score_per is not None and pbr is not None:
                row["PER×PBR\n（15倍以下）"] = float(score_per) * float(pbr)
        except Exception:
            pass

        row["EV/EBITDA（10倍以下）"] = to_float(info.get("enterpriseToEbitda"))

        mc = to_float(info.get("marketCap"))
        if mc is not None:
            row["時価総額\n(百万円)"] = mc / 1_000_000
            row["時価総額_億円"] = mc / 100_000_000

        rg = to_float(info.get("revenueGrowth"))
        eg = to_float(info.get("earningsGrowth"))
        de = to_float(info.get("debtToEquity"))
        if rg is not None:
            row["売上成長率_%"] = rg * 100
        if eg is not None:
            row["利益成長率_%"] = eg * 100
        if de is not None:
            row["負債比率_%"] = de

    except Exception as e:
        row["取得メモ"] = str(e)

    time.sleep(SLEEP)
    return row


def prepare_numeric_columns(df):
    """スコア計算で使う列をすべて安全に数値化する"""
    numeric_cols = [
        "株価",
        "5日線カイリ率(％)",
        "25日線カイリ率(％)",
        "75日線カイリ率(％)",
        "200日線カイリ率(％)",
        "52週高値",
        "52週安値",
        "年間配当_yfinance",
        "配当利回り_%",
        "予想年間配当",
        "PER\n（10倍以下）",
        "PBR\n（1.0倍以下）",
        "PER×PBR\n（15倍以下）",
        "EV/EBITDA（10倍以下）",
        "時価総額\n(百万円)",
        "時価総額_億円",
        "売上成長率_%",
        "利益成長率_%",
        "負債比率_%",
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


# =========================
# 採点ロジック
# =========================
def judge(score_value):
    if score_value >= 90:
        return "A:本命候補"
    elif score_value >= 70:
        return "B:強い候補"
    elif score_value >= 50:
        return "C:監視"
    else:
        return "D:見送り"


def score(df):
    """スコアと判定を作成"""
    df = prepare_numeric_columns(df)

    df["スコア"] = 0
    df["危険減点"] = 0

    # 配当：最大45点
    df.loc[df["配当利回り_%"] >= 3.5, "スコア"] += 25
    df.loc[df["配当利回り_%"] >= 4.5, "スコア"] += 10
    df.loc[df["配当利回り_%"].between(3.5, 6.0, inclusive="both"), "スコア"] += 10

    # バリュー：最大75点
    # PERはyfinanceでズレやすいため、以前より加点を抑制
    df.loc[df["PER\n（10倍以下）"] <= 10, "スコア"] += 15
    df.loc[df["PBR\n（1.0倍以下）"] <= 1.0, "スコア"] += 25
    df.loc[df["PER×PBR\n（15倍以下）"] <= 15, "スコア"] += 10
    df.loc[df["EV/EBITDA（10倍以下）"] <= 10, "スコア"] += 10

    # 安全性・規模：最大15点
    df.loc[df["時価総額\n(百万円)"] >= 100000, "スコア"] += 10
    df.loc[df["時価総額\n(百万円)"] >= 500000, "スコア"] += 5

    # 買いタイミング：最大25点
    df.loc[df["5日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["25日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["75日線カイリ率(％)"] < 0, "スコア"] += 5
    df.loc[df["株価"] < df["52週高値"] * 0.8, "スコア"] += 5
    df.loc[df["株価"] > df["52週安値"] * 1.1, "スコア"] += 5

    # 危険減点
    df.loc[df["配当利回り_%"] > 6, "危険減点"] += 20
    df.loc[df["配当利回り_%"] > 7, "危険減点"] += 25
    df.loc[df["株価"] < df["52週高値"] * 0.6, "危険減点"] += 30
    df.loc[df["75日線カイリ率(％)"] < -10, "危険減点"] += 20
    df.loc[df["PER\n（10倍以下）"] < 5, "危険減点"] += 10
    df.loc[df["売上成長率_%"] < 0, "危険減点"] += 20
    df.loc[df["利益成長率_%"] < 0, "危険減点"] += 20
    df.loc[df["負債比率_%"] > 150, "危険減点"] += 15

    if "33業種" in df.columns:
        df.loc[df["33業種"].isin(["鉄鋼", "海運業", "鉱業"]), "危険減点"] += 10

    df.loc[
        (df["配当利回り_%"] > 4) & (df["PER\n（10倍以下）"] > 20),
        "危険減点"
    ] += 20

    df["総合スコア"] = df["スコア"] - df["危険減点"]
    df["判定"] = df["総合スコア"].apply(judge)

    return df.sort_values(["総合スコア", "配当利回り_%"], ascending=[False, False])


def make_reason(row):
    """高評価銘柄向けの購入推奨理由を作成"""
    if row.get("総合スコア", 0) < 50:
        return ""

    reasons = []

    if row.get("登録銘柄") is True or str(row.get("登録区分", "")) == "登録銘柄":
        reasons.append("登録銘柄")

    dy = row.get("配当利回り_%")
    if pd.notna(dy):
        if dy >= 4.5:
            reasons.append("再計算配当利回り4.5%以上")
        elif dy >= 3.5:
            reasons.append("再計算配当利回り3.5%以上")
        if dy > 6:
            reasons.append("⚠配当高すぎ（減配リスク）")

    per = row.get("PER\n（10倍以下）")
    pbr = row.get("PBR\n（1.0倍以下）")
    per_pbr = row.get("PER×PBR\n（15倍以下）")
    ev_ebitda = row.get("EV/EBITDA（10倍以下）")

    if pd.notna(per) and per <= 10:
        reasons.append("採点用PER10倍以下")
    if pd.notna(per) and per < 5:
        reasons.append("⚠PER低すぎ（要確認）")
    if pd.notna(pbr) and pbr <= 1.0:
        reasons.append("PBR1倍以下")
    if pd.notna(per_pbr) and per_pbr <= 15:
        reasons.append("PER×PBRが15以下")
    if pd.notna(ev_ebitda) and ev_ebitda <= 10:
        reasons.append("EV/EBITDA10倍以下")

    market_cap = row.get("時価総額\n(百万円)")
    if pd.notna(market_cap) and market_cap >= 100000:
        reasons.append("時価総額1000億円以上")

    d25 = row.get("25日線カイリ率(％)")
    d75 = row.get("75日線カイリ率(％)")
    if pd.notna(d25) and d25 < 0:
        reasons.append("25日線下回りで押し目")
    if pd.notna(d75) and d75 < 0:
        reasons.append("75日線下回り")
    if pd.notna(d75) and d75 < -10:
        reasons.append("⚠下落トレンド強い")

    price = row.get("株価")
    high52 = row.get("52週高値")
    low52 = row.get("52週安値")
    if pd.notna(price) and pd.notna(high52) and price < high52 * 0.8:
        reasons.append("52週高値から20%以上下落")
    if pd.notna(price) and pd.notna(high52) and price < high52 * 0.6:
        reasons.append("⚠下がりすぎ")
    if pd.notna(price) and pd.notna(low52) and price > low52 * 1.1:
        reasons.append("52週安値から反発")

    if pd.notna(row.get("売上成長率_%")) and row.get("売上成長率_%") < 0:
        reasons.append("⚠売上減少")
    if pd.notna(row.get("利益成長率_%")) and row.get("利益成長率_%") < 0:
        reasons.append("⚠利益減少")
    if pd.notna(row.get("負債比率_%")) and row.get("負債比率_%") > 150:
        reasons.append("⚠負債比率高い")

    if row.get("33業種") in ["鉄鋼", "海運業", "鉱業"]:
        reasons.append("⚠景気敏感業種")

    # EV/EBITDAのスラッシュでHTML側が分割されないよう、区切りは | にする
    return " | ".join(reasons)


# =========================
# Excel出力
# =========================
def apply_excel_format(writer, df):
    """1行目フィルタ・色・列幅・固定を設定"""
    wb = writer.book

    header_fill = PatternFill("solid", fgColor="9DC3E6")
    header_font = Font(bold=True)
    warning_fill = PatternFill("solid", fgColor="E4DFEC")
    good_fill = PatternFill("solid", fgColor="E2F0D9")
    buy_fill = PatternFill("solid", fgColor="C6E0B4")
    watch_fill = PatternFill("solid", fgColor="FFF2CC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            if any(x in str(cell.value) for x in ["10倍以下", "1.0倍以下", "15倍以下", "EV/EBITDA", "時価総額"]):
                cell.fill = warning_fill

        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)

        ws.row_dimensions[1].height = 36

        header_values = [c.value for c in ws[1]]
        if "判定" in header_values:
            verdict_col = header_values.index("判定") + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                verdict = row[verdict_col - 1].value
                if verdict == "A:本命候補":
                    for c in row:
                        c.fill = buy_fill
                elif verdict == "B:強い候補":
                    for c in row:
                        c.fill = good_fill
                elif verdict == "C:監視":
                    for c in row:
                        c.fill = watch_fill


# =========================
# HTML出力
# =========================
def fmt_html(v, suffix=""):
    if pd.isna(v):
        return "-"
    if isinstance(v, (int, float)):
        return f"{v:,.2f}{suffix}"
    return html.escape(str(v))


def make_link_buttons(code):
    yahoo = f"https://finance.yahoo.co.jp/quote/{code}.T"
    kabutan = f"https://kabutan.jp/stock/?code={code}"
    minkabu = f"https://minkabu.jp/stock/{code}"

    return f"""
      <div class="links">
        <a href="{yahoo}" target="_blank" rel="noopener noreferrer">Yahoo</a>
        <a href="{kabutan}" target="_blank" rel="noopener noreferrer">株探</a>
        <a href="{minkabu}" target="_blank" rel="noopener noreferrer">みんかぶ</a>
      </div>
    """


def make_detail_table(row, columns):
    rows = []
    for col in columns:
        val = row.get(col, "")
        rows.append(f"""
          <tr>
            <th>{html.escape(str(col)).replace(chr(10), "<br>")}</th>
            <td>{fmt_html(val)}</td>
          </tr>
        """)
    return f"""
      <details class="detail-box">
        <summary>全項目を見る</summary>
        <table class="detail-table">
          {''.join(rows)}
        </table>
      </details>
    """


def make_html_report(df, output=OUTPUT_HTML):
    """iPhoneで見やすいHTMLレポートを出力。並び替え・フィルター対応。"""
    display_df = df[df["判定"].isin(["A:本命候補", "B:強い候補", "C:監視"])].copy()
    if display_df.empty:
        display_df = df.head(30).copy()

    now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    cards = []
    detail_columns = list(df.columns)

    markets = sorted([str(x) for x in display_df["市場"].dropna().unique() if str(x).strip()])
    sectors = sorted([str(x) for x in display_df["33業種"].dropna().unique() if str(x).strip()])

    market_options = ''.join(f'<option value="{html.escape(x)}">{html.escape(x)}</option>' for x in markets)
    sector_options = ''.join(f'<option value="{html.escape(x)}">{html.escape(x)}</option>' for x in sectors)

    for _, r in display_df.iterrows():
        verdict = str(r.get("判定", ""))
        cls = "rank-a" if verdict.startswith("A") else "rank-b" if verdict.startswith("B") else "rank-c"

        reason = str(r.get("購入推奨理由", "") or "")
        reason_items = [x.strip() for x in reason.split("|") if x.strip()]
        reason_html = "".join(f"<li>{html.escape(x)}</li>" for x in reason_items)

        per_value = fmt_html(r.get("PER\n（10倍以下）"))
        pbr_value = fmt_html(r.get("PBR\n（1.0倍以下）"))
        per_pbr_value = fmt_html(r.get("PER×PBR\n（15倍以下）"))
        ev_ebitda_value = fmt_html(r.get("EV/EBITDA（10倍以下）"))

        registered_label = "登録銘柄" if str(r.get("登録区分", "")) == "登録銘柄" else "自動抽出"
        registered_class = "registered" if registered_label == "登録銘柄" else "auto-picked"

        code = str(r.get("コード", ""))
        name = str(r.get("銘柄名", ""))
        market = str(r.get("市場", ""))
        sector = str(r.get("33業種", ""))
        text_for_search = f"{code} {name} {market} {sector} {registered_label} {verdict}"

        def data_num(v):
            return "" if pd.isna(v) else str(float(v))

        # f-string内で改行入り列名を直接使うとSyntaxErrorになるため、先に変数化する
        per_data = data_num(r.get("PER\n（10倍以下）"))
        pbr_data = data_num(r.get("PBR\n（1.0倍以下）"))
        perpbr_data = data_num(r.get("PER×PBR\n（15倍以下）"))
        ev_data = data_num(r.get("EV/EBITDA（10倍以下）"))
        marketcap_data = data_num(r.get("時価総額\n(百万円)"))
        dev25_data = data_num(r.get("25日線カイリ率(％)"))
        dev75_data = data_num(r.get("75日線カイリ率(％)"))
        risk_data = data_num(r.get("危険減点"))

        link_buttons = make_link_buttons(code)
        detail_table = make_detail_table(r, detail_columns)

        cards.append(f"""
        <section
          class="card {cls}"
          data-code="{html.escape(code)}"
          data-name="{html.escape(name)}"
          data-market="{html.escape(market)}"
          data-sector="{html.escape(sector)}"
          data-score="{data_num(r.get('総合スコア'))}"
          data-yield="{data_num(r.get('配当利回り_%'))}"
          data-price="{data_num(r.get('株価'))}"
          data-dividend="{data_num(r.get('年間配当_yfinance'))}"
          data-per="{per_data}"
          data-pbr="{pbr_data}"
          data-perpbr="{perpbr_data}"
          data-ev="{ev_data}"
          data-marketcap="{marketcap_data}"
          data-dev25="{dev25_data}"
          data-dev75="{dev75_data}"
          data-risk="{risk_data}"
          data-verdict="{html.escape(verdict)}"
          data-source="{html.escape(registered_label)}"
          data-text="{html.escape(text_for_search)}"
        >
          <div class="topline">
            <div>
              <div class="code">{html.escape(code)} / {html.escape(str(r.get('Ticker', '')))}</div>
              <h2>{html.escape(name)}</h2>
              <div class="sub">{html.escape(market)}｜{html.escape(sector)}｜{html.escape(str(r.get('規模区分','')))}</div>
            </div>
            <div class="badges">
              <div class="badge">{html.escape(verdict)}</div>
              <div class="source-badge {registered_class}">{html.escape(registered_label)}</div>
            </div>
          </div>

          <div class="price-main">
            <span>現在株価</span>
            <strong>{fmt_html(r.get('株価'))}</strong>
          </div>

          {link_buttons}

          <div class="scorebox">
            <div><span>総合スコア</span><strong>{fmt_html(r.get('総合スコア'))}</strong></div>
            <div><span>加点</span><strong>{fmt_html(r.get('スコア'))}</strong></div>
            <div><span>危険減点</span><strong>{fmt_html(r.get('危険減点'))}</strong></div>
          </div>

          <div class="grid">
            <div><span>配当利回り</span><b>{fmt_html(r.get('配当利回り_%'), '%')}</b></div>
            <div><span>年間配当</span><b>{fmt_html(r.get('年間配当_yfinance'))}</b></div>
            <div><span>PER</span><b>{per_value}</b></div>
            <div><span>PBR</span><b>{pbr_value}</b></div>
            <div><span>PER×PBR</span><b>{per_pbr_value}</b></div>
            <div><span>EV/EBITDA</span><b>{ev_ebitda_value}</b></div>
            <div><span>25日乖離</span><b>{fmt_html(r.get('25日線カイリ率(％)'), '%')}</b></div>
            <div><span>75日乖離</span><b>{fmt_html(r.get('75日線カイリ率(％)'), '%')}</b></div>
            <div><span>52週高値</span><b>{fmt_html(r.get('52週高値'))}</b></div>
            <div><span>52週安値</span><b>{fmt_html(r.get('52週安値'))}</b></div>
          </div>

          <div class="reason">
            <h3>購入推奨理由・注意点</h3>
            <ul>{reason_html or '<li>-</li>'}</ul>
          </div>

          {detail_table}
        </section>
        """)

    html_doc = f"""<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>高配当株スクリーニング結果</title>
  <style>
    body {{
      margin: 0;
      background: #f5f6f8;
      font-family: -apple-system, BlinkMacSystemFont, "Helvetica Neue", Arial, sans-serif;
      color: #1f2937;
    }}
    header {{
      position: sticky;
      top: 0;
      background: #111827;
      color: white;
      padding: 14px 16px;
      z-index: 10;
    }}
    header h1 {{
      font-size: 20px;
      margin: 0 0 4px;
    }}
    header p {{
      margin: 0;
      font-size: 12px;
      opacity: .8;
    }}
    main {{
      max-width: 900px;
      margin: 0 auto;
      padding: 12px;
    }}
    .summary, .panel {{
      background: white;
      border-radius: 16px;
      padding: 14px;
      margin-bottom: 12px;
      box-shadow: 0 2px 10px rgba(0,0,0,.06);
    }}
    .summary strong {{
      font-size: 22px;
    }}
    .notice {{
      margin-top: 8px;
      color: #92400e;
      background: #fffbeb;
      border-radius: 10px;
      padding: 8px 10px;
      font-size: 13px;
    }}
    details.panel summary {{
      font-weight: 700;
      cursor: pointer;
      font-size: 16px;
    }}
    .controls {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 8px;
      margin-top: 12px;
    }}
    .controls label {{
      display: flex;
      flex-direction: column;
      font-size: 12px;
      color: #4b5563;
      gap: 3px;
    }}
    .controls input, .controls select {{
      font-size: 16px;
      padding: 9px 10px;
      border-radius: 10px;
      border: 1px solid #d1d5db;
      background: white;
    }}
    @media (max-width: 600px) {{
      .controls {{
        grid-template-columns: 1fr;
      }}
    }}
    .card {{
      background: white;
      border-radius: 18px;
      padding: 16px;
      margin: 12px 0;
      box-shadow: 0 4px 18px rgba(0,0,0,.08);
      border-left: 7px solid #9ca3af;
    }}
    .rank-a {{ border-left-color: #16a34a; }}
    .rank-b {{ border-left-color: #2563eb; }}
    .rank-c {{ border-left-color: #f59e0b; }}
    .topline {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      align-items: flex-start;
    }}
    .code {{ font-size: 12px; color: #6b7280; }}
    h2 {{ margin: 3px 0; font-size: 24px; }}
    .sub {{ font-size: 13px; color: #6b7280; }}
    .badges {{
      display: flex;
      flex-direction: column;
      gap: 6px;
      align-items: flex-end;
    }}
    .badge {{
      white-space: nowrap;
      border-radius: 999px;
      padding: 7px 10px;
      font-size: 12px;
      font-weight: 700;
      background: #eef2ff;
      color: #1e3a8a;
    }}
    .source-badge {{
      white-space: nowrap;
      border-radius: 999px;
      padding: 6px 9px;
      font-size: 11px;
      font-weight: 700;
    }}
    .registered {{ background: #dcfce7; color: #166534; }}
    .auto-picked {{ background: #f3f4f6; color: #4b5563; }}
    .price-main {{
      margin: 14px 0 10px;
      background: linear-gradient(135deg, #eff6ff, #f8fafc);
      border-radius: 16px;
      padding: 14px 16px;
    }}
    .price-main span {{
      display: block;
      font-size: 13px;
      color: #64748b;
      font-weight: 700;
    }}
    .price-main strong {{
      display: block;
      font-size: 42px;
      line-height: 1.1;
      letter-spacing: -1px;
    }}
    .links {{
      display: flex;
      gap: 8px;
      margin: 8px 0 14px;
    }}
    .links a {{
      flex: 1;
      text-align: center;
      text-decoration: none;
      background: #111827;
      color: white;
      padding: 9px 8px;
      border-radius: 10px;
      font-weight: 700;
      font-size: 13px;
    }}
    .scorebox {{
      display: grid;
      grid-template-columns: repeat(3, 1fr);
      gap: 8px;
      margin: 14px 0;
    }}
    .scorebox div {{
      background: #f9fafb;
      border-radius: 12px;
      padding: 10px;
      text-align: center;
    }}
    .scorebox span {{
      display: block;
      font-size: 11px;
      color: #6b7280;
    }}
    .scorebox strong {{ font-size: 22px; }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(2, 1fr);
      gap: 8px;
    }}
    .grid div {{
      background: #f3f4f6;
      border-radius: 12px;
      padding: 9px 10px;
    }}
    .grid span {{
      display: block;
      font-size: 11px;
      color: #6b7280;
    }}
    .grid b {{ font-size: 16px; }}
    .reason {{
      margin-top: 12px;
      background: #fffbeb;
      border-radius: 12px;
      padding: 10px 12px;
    }}
    .reason h3 {{ margin: 0 0 6px; font-size: 14px; }}
    .reason ul {{ margin: 0; padding-left: 20px; }}
    .reason li {{ margin: 3px 0; font-size: 14px; }}
    .detail-box {{
      margin-top: 12px;
      background: #f8fafc;
      border-radius: 12px;
      padding: 10px 12px;
    }}
    .detail-box summary {{
      font-weight: 700;
      cursor: pointer;
    }}
    .detail-table {{
      width: 100%;
      border-collapse: collapse;
      margin-top: 8px;
      font-size: 13px;
    }}
    .detail-table th, .detail-table td {{
      border-bottom: 1px solid #e5e7eb;
      padding: 7px 4px;
      vertical-align: top;
    }}
    .detail-table th {{
      text-align: left;
      width: 44%;
      color: #6b7280;
      font-weight: 700;
    }}
    .hidden {{ display: none; }}
  </style>
</head>
<body>
  <header>
    <h1>高配当株スクリーニング結果</h1>
    <p>作成日時: {now}</p>
  </header>
  <main>
    <div class="summary">
      表示対象: <strong id="visibleCount">{len(display_df)}</strong>件 / 対象表示: {len(display_df)}件 / 全銘柄: {len(df)}件<br>
      A候補: {(df['判定'] == 'A:本命候補').sum()}件　
      B候補: {(df['判定'] == 'B:強い候補').sum()}件　
      C監視: {(df['判定'] == 'C:監視').sum()}件
      <div class="notice">
        配当利回りは「年間配当_yfinance ÷ 株価」で再計算しています。yfinanceのPER・配当・EV/EBITDAは候補抽出用の参考値です。購入前に必ず公式IR・証券会社・株探等で確認してください。
      </div>
    </div>

    <details class="panel">
      <summary>並び替え・フィルターを開く</summary>
      <div class="controls">
        <label>検索（コード・銘柄名・業種）
          <input id="searchInput" type="search" placeholder="例：7272 ヤマハ 銀行">
        </label>
        <label>並び替え
          <select id="sortSelect">
            <option value="score_desc">総合スコア 高い順</option>
            <option value="yield_desc">配当利回り 高い順</option>
            <option value="price_desc">株価 高い順</option>
            <option value="price_asc">株価 低い順</option>
            <option value="dividend_desc">年間配当 高い順</option>
            <option value="per_asc">PER 低い順</option>
            <option value="pbr_asc">PBR 低い順</option>
            <option value="perpbr_asc">PER×PBR 低い順</option>
            <option value="ev_asc">EV/EBITDA 低い順</option>
            <option value="marketcap_desc">時価総額 大きい順</option>
            <option value="dev25_asc">25日乖離 低い順</option>
            <option value="risk_asc">危険減点 少ない順</option>
            <option value="code_asc">コード 昇順</option>
          </select>
        </label>
        <label>判定
          <select id="verdictFilter">
            <option value="">すべて</option>
            <option value="A:本命候補">A:本命候補</option>
            <option value="B:強い候補">B:強い候補</option>
            <option value="C:監視">C:監視</option>
          </select>
        </label>
        <label>登録区分
          <select id="sourceFilter">
            <option value="">すべて</option>
            <option value="登録銘柄">登録銘柄</option>
            <option value="自動抽出">自動抽出</option>
          </select>
        </label>
        <label>市場
          <select id="marketFilter">
            <option value="">すべて</option>
            {market_options}
          </select>
        </label>
        <label>業種
          <select id="sectorFilter">
            <option value="">すべて</option>
            {sector_options}
          </select>
        </label>
        <label>最低スコア
          <input id="minScore" type="number" value="" placeholder="例：70">
        </label>
        <label>最高スコア
          <input id="maxScore" type="number" value="" placeholder="例：150">
        </label>
        <label>最低配当利回り
          <input id="minYield" type="number" step="0.1" value="" placeholder="例：3.5">
        </label>
        <label>最高配当利回り
          <input id="maxYield" type="number" step="0.1" value="" placeholder="例：6.0">
        </label>
        <label>最高PER
          <input id="maxPer" type="number" step="0.1" value="" placeholder="例：12">
        </label>
        <label>最高PBR
          <input id="maxPbr" type="number" step="0.1" value="" placeholder="例：1.0">
        </label>
        <label>最高EV/EBITDA
          <input id="maxEv" type="number" step="0.1" value="" placeholder="例：10">
        </label>
        <label>最低時価総額（百万円）
          <input id="minMarketCap" type="number" value="" placeholder="例：100000">
        </label>
      </div>
    </details>

    <details class="panel">
      <summary>採点基準を見る</summary>
      <h3>加点</h3>
      <ul>
        <li>配当利回りは yfinance値を使わず、年間配当_yfinance ÷ 株価で再計算</li>
        <li>再計算配当利回り 3.5%以上：+25点</li>
        <li>再計算配当利回り 4.5%以上：+10点</li>
        <li>再計算配当利回り 3.5〜6.0%：+10点</li>
        <li>PER 10倍以下：+15点（yfinanceのPERはズレることがあるため控えめ）</li>
        <li>PBR 1.0倍以下：+25点</li>
        <li>PER×PBR 15以下：+10点</li>
        <li>EV/EBITDA 10倍以下：+10点</li>
        <li>時価総額1000億円以上：+10点、5000億円以上：追加+5点</li>
        <li>5日・25日・75日線を下回る：各+5点</li>
        <li>52週高値から20%以上下落：+5点</li>
        <li>52週安値から10%以上反発：+5点</li>
      </ul>
      <h3>危険減点</h3>
      <ul>
        <li>再計算配当利回り 6%超：-20点、7%超：追加-25点</li>
        <li>52週高値から40%以上下落：-30点</li>
        <li>75日線乖離率 -10%未満：-20点</li>
        <li>PER 5倍未満：-10点</li>
        <li>売上成長率マイナス：-20点</li>
        <li>利益成長率マイナス：-20点</li>
        <li>負債比率150%超：-15点</li>
        <li>鉄鋼・海運業・鉱業：-10点</li>
        <li>配当利回り4%超かつPER20倍超：-20点</li>
      </ul>
      <h3>判定</h3>
      <ul>
        <li>90点以上：A 本命候補</li>
        <li>70点以上：B 強い候補</li>
        <li>50点以上：C 監視</li>
        <li>50点未満：D 見送り</li>
      </ul>
    </details>

    <div id="cards">
      {''.join(cards)}
    </div>
  </main>

<script>
  const cardsContainer = document.getElementById('cards');
  const cards = Array.from(document.querySelectorAll('.card'));

  function num(v, fallback) {{
    const n = parseFloat(v);
    return Number.isFinite(n) ? n : fallback;
  }}

  function applyFiltersAndSort() {{
    const q = document.getElementById('searchInput').value.trim().toLowerCase();
    const verdict = document.getElementById('verdictFilter').value;
    const source = document.getElementById('sourceFilter').value;
    const market = document.getElementById('marketFilter').value;
    const sector = document.getElementById('sectorFilter').value;

    const minScore = num(document.getElementById('minScore').value, -Infinity);
    const maxScore = num(document.getElementById('maxScore').value, Infinity);
    const minYield = num(document.getElementById('minYield').value, -Infinity);
    const maxYield = num(document.getElementById('maxYield').value, Infinity);
    const maxPer = num(document.getElementById('maxPer').value, Infinity);
    const maxPbr = num(document.getElementById('maxPbr').value, Infinity);
    const maxEv = num(document.getElementById('maxEv').value, Infinity);
    const minMarketCap = num(document.getElementById('minMarketCap').value, -Infinity);

    const sort = document.getElementById('sortSelect').value;

    let visible = [];

    cards.forEach(card => {{
      const text = (card.dataset.text || '').toLowerCase();
      const score = num(card.dataset.score, -Infinity);
      const yld = num(card.dataset.yield, -Infinity);
      const per = num(card.dataset.per, Infinity);
      const pbr = num(card.dataset.pbr, Infinity);
      const ev = num(card.dataset.ev, Infinity);
      const marketCap = num(card.dataset.marketcap, -Infinity);

      const cardVerdict = card.dataset.verdict || '';
      const cardSource = card.dataset.source || '';
      const cardMarket = card.dataset.market || '';
      const cardSector = card.dataset.sector || '';

      let ok = true;
      if (q && !text.includes(q)) ok = false;
      if (verdict && cardVerdict !== verdict) ok = false;
      if (source && cardSource !== source) ok = false;
      if (market && cardMarket !== market) ok = false;
      if (sector && cardSector !== sector) ok = false;
      if (score < minScore || score > maxScore) ok = false;
      if (yld < minYield || yld > maxYield) ok = false;
      if (per > maxPer) ok = false;
      if (pbr > maxPbr) ok = false;
      if (ev > maxEv) ok = false;
      if (marketCap < minMarketCap) ok = false;

      card.classList.toggle('hidden', !ok);
      if (ok) visible.push(card);
    }});

    visible.sort((a, b) => {{
      if (sort === 'score_desc') return num(b.dataset.score, -Infinity) - num(a.dataset.score, -Infinity);
      if (sort === 'yield_desc') return num(b.dataset.yield, -Infinity) - num(a.dataset.yield, -Infinity);
      if (sort === 'price_desc') return num(b.dataset.price, -Infinity) - num(a.dataset.price, -Infinity);
      if (sort === 'price_asc') return num(a.dataset.price, Infinity) - num(b.dataset.price, Infinity);
      if (sort === 'dividend_desc') return num(b.dataset.dividend, -Infinity) - num(a.dataset.dividend, -Infinity);
      if (sort === 'per_asc') return num(a.dataset.per, Infinity) - num(b.dataset.per, Infinity);
      if (sort === 'pbr_asc') return num(a.dataset.pbr, Infinity) - num(b.dataset.pbr, Infinity);
      if (sort === 'perpbr_asc') return num(a.dataset.perpbr, Infinity) - num(b.dataset.perpbr, Infinity);
      if (sort === 'ev_asc') return num(a.dataset.ev, Infinity) - num(b.dataset.ev, Infinity);
      if (sort === 'marketcap_desc') return num(b.dataset.marketcap, -Infinity) - num(a.dataset.marketcap, -Infinity);
      if (sort === 'dev25_asc') return num(a.dataset.dev25, Infinity) - num(b.dataset.dev25, Infinity);
      if (sort === 'risk_asc') return num(a.dataset.risk, Infinity) - num(b.dataset.risk, Infinity);
      if (sort === 'code_asc') return (a.dataset.code || '').localeCompare(b.dataset.code || '');
      return 0;
    }});

    visible.forEach(card => cardsContainer.appendChild(card));
    document.getElementById('visibleCount').textContent = visible.length;
  }}

  [
    'searchInput', 'sortSelect', 'verdictFilter', 'sourceFilter', 'marketFilter', 'sectorFilter',
    'minScore', 'maxScore', 'minYield', 'maxYield', 'maxPer', 'maxPbr', 'maxEv', 'minMarketCap'
  ].forEach(id => document.getElementById(id).addEventListener('input', applyFiltersAndSort));

  applyFiltersAndSort();
</script>
</body>
</html>"""

    Path(output).write_text(html_doc, encoding="utf-8")
    print("HTML出力完了:", Path(output).resolve())


# =========================
# メイン処理
# =========================
def main():
    master = load_japanese_master()
    codes = build_target_codes(master)

    print(f"{len(codes)}銘柄取得開始")

    rows = []
    for i, code in enumerate(codes, start=1):
        print(f"{i}/{len(codes)}: {code}.T")
        rows.append(fetch(code))

    df = pd.DataFrame(rows)
    df["コード"] = df["コード"].map(normalize_code)

    # sample_codes_japan.csvに入っている銘柄を登録銘柄としてマーク
    registered_codes = set(read_sample_codes(INPUT_CODES))
    df["登録銘柄"] = df["コード"].isin(registered_codes)
    df["登録区分"] = df["登録銘柄"].map(lambda x: "登録銘柄" if x else "自動抽出")

    try:
        df = df.merge(master, on="コード", how="left", suffixes=("", "_master"))

        df["銘柄名"] = df["銘柄名_日本語"].where(
            df["銘柄名_日本語"].notna() & (df["銘柄名_日本語"].astype(str).str.strip() != ""),
            df["銘柄名"]
        )
        df = df.drop(columns=["銘柄名_日本語"])

        for col in ["市場", "33業種", "17業種", "規模区分"]:
            master_col = f"{col}_master"
            if master_col in df.columns:
                df[col] = df[master_col].where(df[master_col].notna(), df[col])
                df = df.drop(columns=[master_col])

        print("日本語銘柄名CSV適用OK")
    except Exception as e:
        print("日本語銘柄名CSVエラー:", e)

    df = score(df)
    df["購入推奨理由"] = df.apply(make_reason, axis=1)
    df = df.round(2)

    preferred_cols = [
        "コード", "銘柄名", "登録区分", "登録銘柄", "市場", "33業種", "17業種", "規模区分", "Ticker",
        "株価",
        "年間配当_yfinance",
        "配当利回り_%",
        "予想年間配当",
        "PER\n（10倍以下）",
        "PBR\n（1.0倍以下）",
        "PER×PBR\n（15倍以下）",
        "EV/EBITDA（10倍以下）",
        "5日線カイリ率(％)",
        "25日線カイリ率(％)",
        "75日線カイリ率(％)",
        "200日線カイリ率(％)",
        "52週高値",
        "52週安値",
        "時価総額\n(百万円)",
        "売上成長率_%",
        "利益成長率_%",
        "負債比率_%",
        "スコア",
        "危険減点",
        "総合スコア",
        "判定",
        "購入推奨理由",
        "取得メモ",
    ]
    cols = [c for c in preferred_cols if c in df.columns] + [c for c in df.columns if c not in preferred_cols]
    df = df[cols]

    output_path = Path(OUTPUT).resolve()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df[df["判定"] == "A:本命候補"].to_excel(writer, index=False, sheet_name="A本命候補")
        df[df["判定"].isin(["A:本命候補", "B:強い候補"])].to_excel(writer, index=False, sheet_name="AB候補")
        df.head(100).to_excel(writer, index=False, sheet_name="ランキング")
        df.to_excel(writer, index=False, sheet_name="全銘柄")
        apply_excel_format(writer, df)

    make_html_report(df, OUTPUT_HTML)

    print("完了:", output_path)
    print("HTML完了:", Path(OUTPUT_HTML).resolve())


if __name__ == "__main__":
    main()
